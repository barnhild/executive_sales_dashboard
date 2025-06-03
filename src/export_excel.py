from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as ExcelImage
from pandas import ExcelWriter
import pandas as pd

def export_report_to_excel(
        path: str,
        monthly_sales: pd.DataFrame,
        region_data: pd.DataFrame,
        product_data: pd.DataFrame,
        forecast_data: pd.DataFrame,
        chart_paths: dict
):
    # Create Excel Workbook
    with ExcelWriter(path, engine='openpyxl') as writer:
        monthly_sales[['Month','Sales']].to_excel(writer, sheet_name='Monthly Summary', index=False)
        region_data.to_excel(writer, sheet_name='Region Breakdown')
        product_data.to_excel(writer, sheet_name='Product Breakdown')
        forecast_data.to_excel(writer, sheet_name='Forecast', index=False)

    print(f"\n Excel summary written to: {path}")

    # reopen and formate in insert images
    wb= load_workbook(path)

    # Bold Header
    ws = wb['Monthly Summary']
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # insert charts
    for sheet_name, image_path in chart_paths.items():
        ws = wb[sheet_name]
        img = ExcelImage(image_path)
        img.anchor = 'F2'
        ws.add_image(img)

    wb.save(path)
    print("Charts embedded into Excel")


